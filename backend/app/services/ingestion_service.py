"""
Ingestion Service — memory-safe document processing with contextual chunking.

Processes PDF, DOCX, TXT, CSV, and XLSX files into vector chunks.
Each chunk is enriched with document metadata (filename, chunk position)
so embeddings capture document-level context, not just raw text.

MEMORY SAFETY: Large PDFs (150MB+) are processed in page batches
to avoid the OOM killer crashing VS Code / uvicorn. Text is never
held entirely in memory — it's extracted, chunked, embedded, and
flushed to DB in streaming fashion.

ACCURACY FEATURES (v2):
  - Heading-aware semantic chunking (splits at ## before fixed-size)
  - Document summary generation (LLM creates a summary stored as chunk #0)
  - Block-based PDF extraction (preserves document structure)
  - Richer contextual chunk headers (file type + position metadata)
"""
import csv
import fitz  # PyMuPDF
import docx
import gc
import os
from typing import List, Optional
from langchain_text_splitters import RecursiveCharacterTextSplitter
from app.interfaces.vector_store import IVectorStore
from app.interfaces.embedder import IEmbedder
from app.interfaces.llm import ILLM
from app.core.logger import logger

# Number of PDF pages to process at a time (kept small for bge-large memory safety).
PDF_PAGE_BATCH_SIZE = 10


class IngestionService:
    def __init__(
        self,
        db: IVectorStore,
        embedder: IEmbedder,
        chunk_size: int = 1000,
        chunk_overlap: int = 300,
        batch_size: int = 50,
        llm: Optional[ILLM] = None,
    ):
        self.db = db
        self.embedder = embedder
        self.batch_size = batch_size
        self.llm = llm  # Optional: used for document summary generation
        # Heading-aware separators — respects document structure before fixed-size splits
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=[
                "\n## ",       # H2 headings (highest priority split)
                "\n### ",      # H3 headings
                "\n#### ",     # H4 headings
                "\n\n",        # Double newline (paragraph boundary)
                "\n",          # Single newline
                ". ",          # Sentence boundary
                " ",           # Word boundary (last resort)
            ],
        )

    def process_file_background(self, file_path: str, filename: str, file_hash: str, tenant_id: str):
        """
        Background worker for processing files of any size.

        Routes by file type:
          PDF  → streaming page-by-page (memory-safe)
          DOCX/TXT/CSV/XLSX → standard processing

        Also generates a document summary (if LLM is available).
        """
        try:
            filename_lower = filename.lower()
            file_type = self._detect_file_type(filename)

            if filename_lower.endswith(".pdf"):
                self._process_pdf_streaming(file_path, filename, file_hash, tenant_id, file_type)
            else:
                self._process_small_file(file_path, filename, file_hash, tenant_id, file_type)

        except Exception as e:
            logger.error(f"Failed to process '{filename}': {e}")
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.debug(f"Cleaned up temp file: {file_path}")

    def _process_pdf_streaming(self, file_path: str, filename: str, file_hash: str, tenant_id: str, file_type: str = "PDF"):
        """Memory-safe PDF processing — pages in batches with block-based extraction."""
        doc = fitz.open(file_path)
        total_pages = doc.page_count
        logger.info(f"Processing PDF '{filename}' | {total_pages} pages | batch size {PDF_PAGE_BATCH_SIZE}")

        total_chunks_saved = 0
        failed_batches = 0
        all_text_for_summary = []  # Collect first pages for summary generation

        for page_start in range(0, total_pages, PDF_PAGE_BATCH_SIZE):
            page_end = min(page_start + PDF_PAGE_BATCH_SIZE, total_pages)

            page_texts = []
            for page_num in range(page_start, page_end):
                page = doc.load_page(page_num)
                # Use 'blocks' extraction to preserve document structure (headings, paragraphs)
                blocks = page.get_text("blocks")
                # Sort blocks by vertical position (top to bottom), then horizontal
                blocks.sort(key=lambda b: (b[1], b[0]))
                page_text = "\n".join(b[4] for b in blocks if b[6] == 0)  # type 0 = text blocks
                page_texts.append(page_text)

            batch_text = "\n".join(page_texts)

            # Collect text from first 3 pages for summary
            if page_start == 0:
                all_text_for_summary.append(batch_text[:3000])

            del page_texts

            if not batch_text.strip():
                continue

            chunks = self.text_splitter.split_text(batch_text)
            del batch_text

            if not chunks:
                continue

            for i in range(0, len(chunks), self.batch_size):
                embed_batch = chunks[i : i + self.batch_size]

                try:
                    contextual_batch = [
                        f"[Document: {filename} | Type: {file_type} | Pages {page_start+1}-{page_end} | Chunk {i+j+1}]\n\n{chunk}"
                        for j, chunk in enumerate(embed_batch)
                    ]
                    embeddings = self.embedder.embed_text(contextual_batch)
                    del contextual_batch

                    records = [
                        {
                            "tenant_id": tenant_id,
                            "filename": filename,
                            "file_hash": file_hash,
                            "content": chunk,
                            "embedding": embeddings[j],
                        }
                        for j, chunk in enumerate(embed_batch)
                    ]

                    self.db.save_documents(records)
                    total_chunks_saved += len(embed_batch)
                    del records, embeddings

                except Exception as e:
                    failed_batches += 1
                    logger.error(f"Failed batch for pages {page_start+1}-{page_end}: {e}")

            del chunks
            gc.collect()

            logger.info(f"PDF '{filename}' | pages {page_start+1}-{page_end}/{total_pages} | {total_chunks_saved} chunks")

        doc.close()

        # Generate and store document summary as chunk #0
        if all_text_for_summary:
            self._generate_document_summary(
                text_preview="\n".join(all_text_for_summary),
                filename=filename, file_hash=file_hash, tenant_id=tenant_id, file_type=file_type,
            )

        if failed_batches > 0:
            logger.warning(f"Completed '{filename}' with {failed_batches} failed batches | {total_chunks_saved} chunks")
        else:
            logger.info(f"Successfully ingested '{filename}' | {total_pages} pages → {total_chunks_saved} chunks")

    def _process_small_file(self, file_path: str, filename: str, file_hash: str, tenant_id: str, file_type: str = "Document"):
        """Standard processing for DOCX, TXT, CSV, and XLSX files."""
        raw_text = self._extract_text_from_disk(file_path, filename)
        logger.info(f"Extracted text from '{filename}' ({len(raw_text)} chars)")

        if not raw_text.strip():
            raise ValueError(f"No readable text was found in '{filename}'")

        chunks = self.text_splitter.split_text(raw_text)
        total_chunks = len(chunks)
        logger.info(f"Split '{filename}' into {total_chunks} chunks")

        # Generate document summary before deleting raw_text
        self._generate_document_summary(
            text_preview=raw_text[:3000],
            filename=filename, file_hash=file_hash, tenant_id=tenant_id, file_type=file_type,
        )
        del raw_text

        total_batches = (total_chunks + self.batch_size - 1) // self.batch_size
        failed_batches = 0

        for i in range(0, total_chunks, self.batch_size):
            batch_num = i // self.batch_size + 1
            batch_chunks = chunks[i : i + self.batch_size]

            try:
                contextual_batch = [
                    f"[Document: {filename} | Type: {file_type} | Chunk {i + j + 1}/{total_chunks}]\n\n{chunk}"
                    for j, chunk in enumerate(batch_chunks)
                ]

                embeddings = self.embedder.embed_text(contextual_batch)
                del contextual_batch

                records = [
                    {
                        "tenant_id": tenant_id,
                        "filename": filename,
                        "file_hash": file_hash,
                        "content": chunk,
                        "embedding": embeddings[j],
                    }
                    for j, chunk in enumerate(batch_chunks)
                ]

                self.db.save_documents(records)
                del records, embeddings
                logger.info(f"Processed batch {batch_num}/{total_batches} for '{filename}'")

            except Exception as e:
                failed_batches += 1
                logger.error(f"Failed batch {batch_num}/{total_batches} for '{filename}': {e}")

        if failed_batches > 0:
            logger.warning(f"Completed '{filename}' with {failed_batches}/{total_batches} failed batches")
        else:
            logger.info(f"Successfully ingested '{filename}' ({total_batches} batches)")

    def delete_file(self, filename: str, tenant_id: str) -> bool:
        """Deletes all chunks of a document."""
        return self.db.delete_document(filename=filename, tenant_id=tenant_id)

    def list_files(self, tenant_id: str) -> List[str]:
        """Lists all unique filenames for a tenant."""
        return self.db.get_all_documents(tenant_id=tenant_id)

    def _extract_text_from_disk(self, file_path: str, filename: str) -> str:
        """
        Extracts readable text from non-PDF files.

        Supports: DOCX, TXT, CSV, XLSX
        PDFs use the streaming method instead.
        """
        filename_lower = filename.lower()

        if filename_lower.endswith(".docx"):
            doc_file = docx.Document(file_path)
            return "\n".join([paragraph.text for paragraph in doc_file.paragraphs])

        elif filename_lower.endswith(".txt"):
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()

        elif filename_lower.endswith(".csv"):
            return self._extract_csv(file_path)

        elif filename_lower.endswith((".xlsx", ".xls")):
            return self._extract_xlsx(file_path)

        else:
            raise ValueError(f"Unsupported file type: {filename}")

    def _extract_csv(self, file_path: str) -> str:
        """
        Converts CSV into readable text format.

        Each row becomes: "Column1: value1 | Column2: value2 | ..."
        This makes the data semantically searchable.
        """
        rows = []
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row_text = " | ".join(
                        f"{key}: {value}" for key, value in row.items()
                        if value and value.strip()
                    )
                    if row_text:
                        rows.append(row_text)
        except Exception as e:
            logger.warning(f"CSV DictReader failed, falling back to raw read: {e}")
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()

        return "\n".join(rows)

    def _extract_xlsx(self, file_path: str) -> str:
        """
        Extracts text from Excel files (XLSX/XLS).

        Reads all sheets, converting each row to "Col: val | Col: val" format.
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            all_text = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_text.append(f"--- Sheet: {sheet_name} ---")

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # Use first row as headers
                headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]

                for row in rows[1:]:
                    row_text = " | ".join(
                        f"{headers[i]}: {str(cell)}"
                        for i, cell in enumerate(row)
                        if cell is not None and str(cell).strip()
                    )
                    if row_text:
                        all_text.append(row_text)

            wb.close()
            return "\n".join(all_text)

        except ImportError:
            logger.warning("openpyxl not installed, cannot process XLSX files")
            raise ValueError("XLSX support requires openpyxl: pip install openpyxl")
        except Exception as e:
            logger.error(f"XLSX extraction failed: {e}")
            raise

    # ── New: Accuracy Enhancement Helpers ──

    def _detect_file_type(self, filename: str) -> str:
        """Returns a human-readable file type label for contextual headers."""
        ext = os.path.splitext(filename)[1].lower()
        type_map = {
            ".pdf": "PDF", ".docx": "Word Document", ".txt": "Text File",
            ".csv": "CSV Spreadsheet", ".xlsx": "Excel Spreadsheet",
            ".xls": "Excel Spreadsheet",
        }
        return type_map.get(ext, "Document")

    def _generate_document_summary(self, text_preview: str, filename: str, file_hash: str,
                                    tenant_id: str, file_type: str):
        """
        Uses the LLM to generate a document summary and stores it as a special chunk.

        This helps answer "what is this document about?" queries and improves
        retrieval for broad questions about document contents.
        """
        if not self.llm:
            return  # No LLM available — skip summary generation

        try:
            summary = self.llm.generate_response(
                system_prompt=(
                    "You are a document summarizer. Given the beginning of a document, "
                    "write a concise 3-5 sentence summary describing what the document contains, "
                    "its key topics, and its purpose. Be factual and specific."
                ),
                user_prompt=f"Document: {filename} (Type: {file_type})\n\nContent preview:\n{text_preview[:2500]}",
                temperature=0.0,
            )

            if not summary or len(summary) < 20:
                logger.warning(f"Summary generation returned empty result for '{filename}'")
                return

            # Store summary as a special chunk with [SUMMARY] prefix
            summary_content = f"[DOCUMENT SUMMARY] {filename}\n\n{summary}"
            contextual_text = f"[Document: {filename} | Type: {file_type} | Summary]\n\n{summary_content}"
            summary_embedding = self.embedder.embed_text([contextual_text])[0]

            self.db.save_documents([{
                "tenant_id": tenant_id,
                "filename": filename,
                "file_hash": file_hash,
                "content": summary_content,
                "embedding": summary_embedding,
            }])

            logger.info(f"Generated and stored document summary for '{filename}'")

        except Exception as e:
            logger.warning(f"Document summary generation failed for '{filename}': {e}")
